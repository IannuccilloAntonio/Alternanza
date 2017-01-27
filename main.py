from openpyxl import load_workbook
import optparse

parser = optparse.OptionParser('usage: -s -e -v')
parser.add_option('-s', dest='colIniz', type='string', help='colonna iniziale')
parser.add_option('-e', dest='colFin', type='string', help='colonna finale')
parser.add_option('-v', dest='colVal', type='string', help='colonna valori')
parser.add_option('-r', dest='rowStart', type='string', help='riga iniziale')
(options, args) = parser.parse_args()
if options.colIniz==None or options.colFin==None or options.colVal==None or options.rowStart==None:
    print parser.usage
    exit(0)

wb = load_workbook(filename = 'test.xlsx')
ws = wb.active

colIniz = options.colIniz
colFin = options.colFin
colVal = options.colVal

row = int(options.rowStart)
while ws[colIniz+str(row)].value != None:
    print ws[colIniz+str(row)].value
    if ws[colIniz+str(row)].value == ws[colFin+str(row)].value:
        ws[colIniz+str(row)].value = ws[colVal+str(row)].value
    row += 1

wb.save(filename = 'test2.xlsx')
